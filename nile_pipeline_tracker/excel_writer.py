from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .config import CATEGORIES, OTHER_CATEGORY, ALL_CATS

# Column layout of Raw Tickets sheet — used to build formula references
_RAW_COLS = [
    "Key", "Issue id", "Parent id", "Summary", "Description", "Comment", "Close Notes",
    "Categories", "New Categories", "Assign", "Status", "Created", "Updated",
    "Custom field (Assignment Group)", "Assignee", "Reporter", "Resolution",
    "Custom field (Comments)", "Inward issue link (Cloners)", "Outward issue link (Cloners)",
    "Inward issue link (Relates)", "Category", "Review Required",
]
_RAW_CAT_COL = get_column_letter(_RAW_COLS.index("Category") + 1)  # "V"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="F2F7FF")
THIN = Side(style="thin", color="CCCCCC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CAT_COLOR = {c["name"]: c["color"] for c in CATEGORIES}
CAT_COLOR[OTHER_CATEGORY["name"]] = OTHER_CATEGORY["color"]


def load_prev_month_counts(output_dir: str) -> dict:
    """
    Scan output/ for the most recent CSV report from a prior month.
    Returns a dict of {category_name: count} or empty dict if none found.
    """
    import csv as _csv
    import glob
    import os
    import time

    pattern   = os.path.join(output_dir, "Nile_Pipeline_Issue_Count_*.csv")
    files     = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)

    if not files:
        return {}

    now       = time.time()
    prev_file = None
    for f in files:
        if now - os.path.getmtime(f) > 600:
            prev_file = f
            break

    if not prev_file:
        return {}

    counts = {}
    try:
        with open(prev_file, newline="", encoding="utf-8") as f:
            reader = _csv.DictReader(f)
            for row in reader:
                cat = row.get("Category", "").strip()
                if cat:
                    counts[cat] = counts.get(cat, 0) + 1
    except Exception:
        return {}

    return counts


def write_cover_sheet(ws, rows: list[dict], month_label: str,
                      cat_counts: dict, prev_counts: dict,
                      raw_sheet_name: str = "Raw Tickets"):
    """
    Executive cover sheet — the first sheet the report opens on.
    Contains: branded header, 4 KPI cards, health status,
    category snapshot table with trend arrows and mini bars.
    All data comes from rows/cat_counts — no manual input needed.
    """
    ws.title = "Cover Sheet"

    total       = len(rows)
    other_count = cat_counts.get(OTHER_CATEGORY["name"], 0)
    other_pct   = round(other_count / total * 100, 1) if total else 0

    # Find top 2 categories (excluding Other)
    issue_cats = [(n, c) for n, c in cat_counts.items()
                  if n != OTHER_CATEGORY["name"] and c > 0]
    issue_cats.sort(key=lambda x: x[1], reverse=True)
    top1_name  = issue_cats[0][0] if issue_cats else "N/A"
    top1_count = issue_cats[0][1] if issue_cats else 0
    top2_name  = issue_cats[1][0] if len(issue_cats) > 1 else ""
    top2_count = issue_cats[1][1] if len(issue_cats) > 1 else 0

    # ── Branded header rows 1-3 ────────────────────────────────
    NAVY  = "1F3864"
    WHITE = "FFFFFF"
    LIGHT = "B5D4F4"

    ws.merge_cells("A1:H1")
    h1 = ws["A1"]
    h1.value     = "NILE Platform — Production Support"
    h1.font      = Font(bold=True, size=16, color=WHITE)
    h1.fill      = PatternFill("solid", fgColor=NAVY)
    h1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    h2 = ws["A2"]
    h2.value     = f"Monthly Issue Report  ·  {month_label}  ·  Auto-generated"
    h2.font      = Font(size=11, color=LIGHT)
    h2.fill      = PatternFill("solid", fgColor=NAVY)
    h2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:H3")
    ws["A3"].fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[3].height = 8

    # ── 4 KPI cards (rows 5-6) ─────────────────────────────────
    kpis = [
        ("Total Issues",   str(total),       "E6F1FB", "0C447C", "185FA5"),
        (top1_name,        str(top1_count),  "FCEBEB", "791F1F", "A32D2D"),
        (top2_name or "—", str(top2_count),  "FAEEDA", "633806", "854F0B"),
        ("Needs Review",   f"{other_pct}%",  "FCEBEB", "791F1F", "A32D2D"),
    ]

    ws.row_dimensions[4].height = 10
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 32
    ws.row_dimensions[7].height = 10

    for ki, (label, value, bg, txt_dark, txt_mid) in enumerate(kpis):
        col_start    = ki * 2 + 1
        col_end      = col_start + 1
        col_letter_s = get_column_letter(col_start)
        col_letter_e = get_column_letter(col_end)

        ws.merge_cells(f"{col_letter_s}5:{col_letter_e}5")
        lc = ws[f"{col_letter_s}5"]
        lc.value     = label
        lc.font      = Font(size=10, color=txt_mid)
        lc.fill      = PatternFill("solid", fgColor=bg)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = Border(
            left=Side(style="thin", color=txt_mid),
            right=Side(style="thin", color=txt_mid),
            top=Side(style="thin", color=txt_mid),
        )

        ws.merge_cells(f"{col_letter_s}6:{col_letter_e}6")
        vc = ws[f"{col_letter_s}6"]
        _cat_range = f"'{raw_sheet_name}'!${_RAW_CAT_COL}$3:${_RAW_CAT_COL}$10000"
        if ki == 0:  # Total Issues
            vc.value = f'=COUNTIF({_cat_range},"<>")'
        elif ki == 3:  # Needs Review %
            _total = f'COUNTIF({_cat_range},"<>")'
            _other = f'COUNTIF({_cat_range},"Other")'
            vc.value = f'=TEXT(IF({_total}>0,{_other}/{_total},0),"0.0%")'
        else:
            vc.value = value
        vc.font      = Font(bold=True, size=22, color=txt_dark)
        vc.fill      = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = Border(
            left=Side(style="thin", color=txt_mid),
            right=Side(style="thin", color=txt_mid),
            bottom=Side(style="thin", color=txt_mid),
        )

    # ── Health status row 9 ────────────────────────────────────
    ws.row_dimensions[8].height  = 10
    ws.row_dimensions[9].height  = 28
    ws.row_dimensions[10].height = 10

    if other_pct > 40:
        health_bg   = "FFF3CD"
        health_txt  = "856404"
        health_icon = "⚠"
        health_msg  = (
            f"Needs Review rate is high ({other_pct}%). "
            f"Top issue: {top1_name} ({top1_count}). "
            "Consider expanding keywords in config.py."
        )
    else:
        health_bg   = "D4EDDA"
        health_txt  = "155724"
        health_icon = "✓"
        health_msg  = (
            f"Categorization rate is healthy. "
            f"Top issue this month: {top1_name} ({top1_count}). "
            f"Total issues tracked: {total}."
        )

    ws.merge_cells("A9:H9")
    hc = ws["A9"]
    hc.value     = f"{health_icon}  Health Status:  {health_msg}"
    hc.font      = Font(size=11, color=health_txt, bold=True)
    hc.fill      = PatternFill("solid", fgColor=health_bg)
    hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hc.border    = Border(
        left=Side(style="thin", color=health_txt),
        right=Side(style="thin", color=health_txt),
        top=Side(style="thin", color=health_txt),
        bottom=Side(style="thin", color=health_txt),
    )

    # ── Category snapshot table from row 12 ────────────────────
    ws.row_dimensions[11].height = 10

    snap_hdr_fill = PatternFill("solid", fgColor=NAVY)
    snap_hdrs     = ["Category", "Count", "% Share", "vs Last Month", "Trend Bar"]

    ws.row_dimensions[12].height = 22
    for ci, h in enumerate(snap_hdrs, 1):
        c = ws.cell(row=12, column=ci, value=h)
        c.fill      = snap_hdr_fill
        c.font      = Font(bold=True, size=11, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BOX

    sorted_cats = [(n, cat_counts[n]) for n in ALL_CATS if cat_counts.get(n, 0) > 0]
    max_count   = max((c for _, c in sorted_cats), default=1)

    for ri, (cat_name, count) in enumerate(sorted_cats, start=13):
        pct      = round(count / total * 100, 1) if total else 0
        color    = CAT_COLOR.get(cat_name, "D9D9D9")
        alt      = ri % 2 == 0
        alt_fill = PatternFill("solid", fgColor="F2F7FF") if alt else None

        prev = prev_counts.get(cat_name, 0)
        diff = count - prev
        if prev == 0 and count > 0:
            vs_txt  = "New"
            vs_font = Font(size=11, color="185FA5", bold=True)
        elif diff > 0:
            vs_txt  = f"▲ +{diff}"
            vs_font = Font(size=11, color="3B6D11", bold=True)
        elif diff < 0:
            vs_txt  = f"▼ {diff}"
            vs_font = Font(size=11, color="A32D2D", bold=True)
        else:
            vs_txt  = "— same"
            vs_font = Font(size=11, color="888780")

        bar_len = max(1, round((count / max_count) * 20))
        bar_txt = "█" * bar_len

        row_data = [cat_name, count, f"{pct}%", vs_txt, bar_txt]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = BOX
            if alt_fill:
                c.fill = alt_fill
            if ci == 1:
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif ci == 5:
                c.font      = Font(size=10, color=color, bold=True)
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif ci == 4:
                c.font      = vs_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
                if ci == 2:
                    c.font = Font(bold=True, size=11)

        ws.row_dimensions[ri].height = 20

    # Total row
    total_ri = 13 + len(sorted_cats)
    tot_fill = PatternFill("solid", fgColor="E8EDF5")
    for ci in range(1, 6):
        c = ws.cell(row=total_ri, column=ci)
        c.fill      = tot_fill
        c.border    = BOX
        c.font      = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_ri, column=1, value="TOTAL").alignment = \
        Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=total_ri, column=2, value=total)
    ws.cell(row=total_ri, column=3, value="100%")
    ws.row_dimensions[total_ri].height = 22

    # Footer note
    footer_row = total_ri + 2
    ws.merge_cells(f"A{footer_row}:H{footer_row}")
    fc = ws[f"A{footer_row}"]
    fc.value = (
        "All figures are auto-generated from the Jira export. "
        "Re-run the tracker to refresh. "
        "Needs Review tickets require manual categorization."
    )
    fc.font      = Font(size=9, color="888780", italic=True)
    fc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[footer_row].height = 16

    # Column widths
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"


def write_raw_sheet(ws, rows: list[dict], month_label: str):
    ws.title = "Raw Tickets"
    cols = [
        "Key", "Issue id", "Parent id", "Summary", "Description", "Comment", "Close Notes",
        "Categories", "New Categories", "Assign", "Status", "Created", "Updated",
        "Custom field (Assignment Group)", "Assignee", "Reporter", "Resolution",
        "Custom field (Comments)", "Inward issue link (Cloners)", "Outward issue link (Cloners)",
        "Inward issue link (Relates)", "Category", "Review Required",
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    title = ws["A1"]
    title.value = f"NILE Pipeline Issues — {month_label}"
    title.font = Font(bold=True, size=14, color="1F3864")
    title.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    for ci, col_name in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=col_name)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX
    ws.row_dimensions[2].height = 22

    for ri, row in enumerate(rows, 3):
        for ci, col_name in enumerate(cols, 1):
            val = row.get(col_name, "")
            wrap = col_name in {"Summary", "Description", "Comment", "Close Notes", "Custom field (Comments)"}
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(wrap_text=wrap, vertical="top")
            c.border = BOX
            if ri % 2 == 0:
                c.fill = ALT_FILL
            if col_name == "Category" and val:
                cat_color = CAT_COLOR.get(val, "")
                if cat_color:
                    c.font = Font(color=cat_color, bold=True, size=10)

    widths = [12, 12, 12, 42, 42, 30, 24, 20, 20, 18, 14, 14, 14, 24, 18, 18, 24, 20, 20, 20, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"

    # ── Category dropdown validation ───────────────────────────
    category_col_index = cols.index("Category") + 1
    cat_col_letter = get_column_letter(category_col_index)

    wb = ws.parent
    if "_categories" not in wb.sheetnames:
        cat_ws = wb.create_sheet("_categories")
        for i, cat in enumerate(ALL_CATS, start=1):
            cat_ws.cell(row=i, column=1, value=cat)
        cat_ws.sheet_state = "hidden"
    else:
        cat_ws = wb["_categories"]

    cat_range = f"'_categories'!$A$1:$A${len(ALL_CATS)}"
    dv = DataValidation(
        type="list",
        formula1=cat_range,
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Category",
        error="Please select a valid category from the dropdown.",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"{cat_col_letter}3:{cat_col_letter}{2 + len(rows)}"


def write_summary_sheet(ws, rows: list[dict], month_label: str,
                        prev_counts: dict = None,
                        raw_sheet_name: str = "Raw Tickets"):
    ws.title = "Monthly Summary"
    counts = defaultdict(int)
    for row in rows:
        counts[row["Category"]] += 1

    sorted_cats = [(name, counts[name]) for name in ALL_CATS]

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = f"NILE Pipeline Issue Summary — {month_label}"
    title.font = Font(bold=True, size=16, color="1F3864")
    title.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    cat_ref = f"'{raw_sheet_name}'!${_RAW_CAT_COL}$3:${_RAW_CAT_COL}$10000"
    ws["A2"].value = f'="Total Tickets: "&TEXT(COUNTIF({cat_ref},"<>"),"0")'
    ws["A2"].font = Font(bold=True, size=12, color="444444")
    ws.row_dimensions[2].height = 20

    headings = ["Category", "Count", "% Share", "vs Last Month", "Color Tag"]
    for ci, heading in enumerate(headings, 1):
        c = ws.cell(row=4, column=ci, value=heading)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX
    ws.row_dimensions[4].height = 22

    first_data_row = 5
    for ri, (cat_name, count) in enumerate(sorted_cats, start=first_data_row):
        color = CAT_COLOR.get(cat_name, "D9D9D9")
        alt   = ri % 2 == 0

        c1 = ws.cell(row=ri, column=1, value=cat_name)
        c1.alignment = Alignment(horizontal="left")
        c1.border = BOX
        if alt:
            c1.fill = ALT_FILL

        countif_expr = f'COUNTIF({cat_ref},"{cat_name}")'
        c2 = ws.cell(row=ri, column=2, value=f"={countif_expr}")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.font = Font(bold=True, size=11)
        c2.border = BOX
        if alt:
            c2.fill = ALT_FILL

        total_expr = f"COUNTIF({cat_ref},\"<>\")"
        c3 = ws.cell(row=ri, column=3,
                     value=f"=IF({total_expr}>0,{countif_expr}/{total_expr},0)")
        c3.number_format = "0.0%"
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c3.border = BOX
        if alt:
            c3.fill = ALT_FILL

        # vs Last Month (column 4)
        prev     = (prev_counts or {}).get(cat_name, 0)
        diff     = count - prev
        if prev == 0 and count > 0:
            vs_txt   = "New"
            vs_color = "185FA5"
        elif diff > 0:
            vs_txt   = f"▲ +{diff}"
            vs_color = "3B6D11"
        elif diff < 0:
            vs_txt   = f"▼ {diff}"
            vs_color = "A32D2D"
        else:
            vs_txt   = "— same"
            vs_color = "888780"

        c4 = ws.cell(row=ri, column=4, value=vs_txt)
        c4.font      = Font(bold=True, size=11, color=vs_color)
        c4.alignment = Alignment(horizontal="center", vertical="center")
        c4.border    = BOX
        if alt:
            c4.fill = ALT_FILL

        # Color Tag moved to column 5
        color_cell = ws.cell(row=ri, column=5, value="")
        color_cell.fill   = PatternFill("solid", fgColor=color)
        color_cell.border = BOX

    total_row = first_data_row + len(sorted_cats)
    if total_row > first_data_row:
        last_data_row = total_row - 1
        t1 = ws.cell(row=total_row, column=1, value="Monthly Total")
        t1.font   = Font(bold=True)
        t1.border = BOX

        t2 = ws.cell(row=total_row, column=2,
                     value=f"=SUM(B{first_data_row}:B{last_data_row})")
        t2.font      = Font(bold=True, size=11)
        t2.alignment = Alignment(horizontal="center", vertical="center")
        t2.border    = BOX

        t3 = ws.cell(row=total_row, column=3, value="100%")
        t3.alignment = Alignment(horizontal="center")
        t3.border    = BOX

        t4 = ws.cell(row=total_row, column=4, value="")
        t4.border = BOX

        t5 = ws.cell(row=total_row, column=5, value="")
        t5.border = BOX

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    chart_end = 4 + len(sorted_cats)
    if sorted_cats:
        bar = BarChart()
        bar.type = "col"
        bar.title = f"Issue Count by Category — {month_label}"
        bar.y_axis.title = "Count"
        bar.style = 10
        bar.width = 26
        bar.height = 15

        data = Reference(ws, min_col=2, min_row=4, max_row=chart_end)
        cats = Reference(ws, min_col=1, min_row=5, max_row=chart_end)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)

        for idx, (cat_name, _) in enumerate(sorted_cats):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = CAT_COLOR.get(cat_name, "D9D9D9")
            bar.series[0].dPt.append(pt)

        ws.add_chart(bar, "F4")

        pie = PieChart()
        pie.title = "Issue Distribution"
        pie.style = 10
        pie.width = 18
        pie.height = 14

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        for idx, (cat_name, _) in enumerate(sorted_cats):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = CAT_COLOR.get(cat_name, "D9D9D9")
            pie.series[0].dPt.append(pt)

        ws.add_chart(pie, "F24")

    ws.freeze_panes = "A5"


def write_uncategorized_sheet(ws, rows: list[dict],
                               raw_sheet_name: str = "Raw Tickets"):
    ws.title = "Needs Review"

    other_rows  = [r for r in rows if r.get("Category") == "Other"]
    total_other = len(other_rows)

    # ── Title row ─────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    h1 = ws["A1"]
    h1.value     = f"Needs Review — {total_other} uncategorized tickets"
    h1.font      = Font(bold=True, size=13, color="FFFFFF")
    h1.fill      = PatternFill("solid", fgColor="C00000")
    h1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Live count linked to Raw Tickets ──────────────────────
    cat_range = f"'{raw_sheet_name}'!${_RAW_CAT_COL}$3:${_RAW_CAT_COL}$10000"
    ws.merge_cells("A2:G2")
    c2 = ws["A2"]
    c2.value = (
        f"=\"Needs Review count: \""
        f"&COUNTIF({cat_range},\"Other\")"
        "&\" — Change category in Raw Tickets tab, then re-run to refresh this list\""
    )
    c2.font      = Font(size=11, italic=True, color="C00000")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # ── Column headers ────────────────────────────────────────
    hdrs = ["Key", "Summary", "Description", "Comment", "Close Notes", "Status", "Created"]
    hdr_fill = PatternFill("solid", fgColor="C00000")
    ws.row_dimensions[3].height = 20
    for ci, hdr in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=ci, value=hdr)
        c.fill      = hdr_fill
        c.font      = Font(bold=True, size=11, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BOX

    # ── Static data rows ──────────────────────────────────────
    for ri, row in enumerate(other_rows, start=4):
        alt      = ri % 2 == 0
        alt_fill = PatternFill("solid", fgColor="FFF2F2") if alt else None
        vals = [
            row.get("key",         row.get("Key",         "")),
            row.get("summary",     row.get("Summary",     "")),
            row.get("description", row.get("Description", "")),
            row.get("comments",    row.get("Comment",     "")),
            row.get("close_notes", row.get("Close Notes", "")),
            row.get("status",      row.get("Status",      "")),
            row.get("created",     row.get("Created",     "")),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val or "")
            c.border = BOX
            c.font   = Font(size=10)
            if alt_fill:
                c.fill = alt_fill
            if ci in (2, 3, 4, 5):
                c.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[ri].height = 40
            else:
                c.alignment = Alignment(vertical="top")

    # ── Footer note ───────────────────────────────────────────
    note_row = 4 + total_other + 1
    ws.merge_cells(f"A{note_row}:G{note_row}")
    note = ws[f"A{note_row}"]
    note.value = (
        f"To recategorize: go to {raw_sheet_name} → find ticket → "
        "use Category dropdown → re-run python run.py to refresh this list."
    )
    note.font      = Font(size=9, italic=True, color="888780")
    note.alignment = Alignment(horizontal="center")

    # ── Column widths ─────────────────────────────────────────
    widths = [14, 42, 38, 32, 28, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ── Multi-month helpers ───────────────────────────────────────────────────────

def _short_label(month_label: str) -> str:
    """Convert 'April 2026' → 'Apr 2026', 'May 2026' → 'May 2026'."""
    parts = month_label.strip().split()
    if len(parts) >= 2:
        month_name = parts[0]
        year       = parts[-1]
        if len(month_name) > 3:
            month_name = month_name[:3]
        return f"{month_name} {year}"
    return month_label[:10]


def write_month_raw_sheet(wb, rows: list[dict], month_label: str):
    """
    Write raw tickets for one month as a named sheet.
    Sheet name: 'Raw — Apr 2026', 'Raw — May 2026' etc.
    If the sheet already exists, skip (never overwrite previous month).
    """
    short      = _short_label(month_label)
    sheet_name = f"Raw — {short}"

    if sheet_name in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' already exists — skipping.")
        return wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    write_raw_sheet(ws, rows, month_label)
    ws.title = sheet_name   # override the title set inside write_raw_sheet
    return ws


def write_all_months_summary(wb, all_months_data: list[dict]):
    """
    Write or fully refresh the 'All Months Summary' sheet.
    all_months_data = list of dicts, one per month:
      {"label": "Apr 2026", "counts": {"Configuration Issue": 31, ...}}
    Sorted oldest → newest.

    Layout: Category | Count(Apr) | %(Apr) | Count(May) | %(May) | ...
    All counts are COUNTIF formulas referencing each Raw sheet.
    Totals use SUM formulas.  Chart is horizontal bars with data labels.
    """
    SHEET_NAME = "All Months Summary"
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    # Insert at position 1 (right after Cover Sheet which is always position 0)
    ws = wb.create_sheet(SHEET_NAME, 1)

    NAVY  = "1F3864"
    WHITE = "FFFFFF"

    # 2 columns per month: Count + % Share
    # Total columns = 1 (Category) + 2 * n_months
    n_months        = len(all_months_data)
    last_col_num    = 1 + 2 * max(n_months, 1)
    last_col_letter = get_column_letter(last_col_num)

    # ── Title (row 1) ───────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col_letter}1")
    t = ws["A1"]
    t.value     = "NILE Platform — Production Support  |  All Months Issue Summary"
    t.font      = Font(bold=True, size=14, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Column headers (row 3) ──────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    hdr_fill = PatternFill("solid", fgColor=NAVY)

    h0 = ws.cell(row=3, column=1, value="Category")
    h0.fill      = hdr_fill
    h0.font      = Font(bold=True, size=11, color=WHITE)
    h0.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    h0.border    = BOX

    col = 2
    for month in all_months_data:
        hc = ws.cell(row=3, column=col, value=month["label"])
        hc.fill      = hdr_fill
        hc.font      = Font(bold=True, size=11, color=WHITE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border    = BOX

        hp = ws.cell(row=3, column=col + 1, value="% Share")
        hp.fill      = hdr_fill
        hp.font      = Font(bold=True, size=10, color=WHITE)
        hp.alignment = Alignment(horizontal="center", vertical="center")
        hp.border    = BOX
        col += 2

    # ── Data rows — COUNTIF formulas referencing each Raw sheet ─────────
    # total_row pre-computed so % formula can reference it
    total_row = 4 + len(ALL_CATS)

    for ri, cat_name in enumerate(ALL_CATS, 4):
        alt      = ri % 2 == 0
        alt_fill = PatternFill("solid", fgColor="F2F7FF") if alt else None

        c1 = ws.cell(row=ri, column=1, value=cat_name)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c1.border    = BOX
        if alt_fill:
            c1.fill = alt_fill

        col = 2
        for month in all_months_data:
            raw_tab  = f"Raw — {month['label']}"
            safe_tab = raw_tab.replace("'", "''")

            # Count — live COUNTIF against the Raw sheet
            count_formula = (
                f"=COUNTIF('{safe_tab}'!$V$3:$V$10000,\"{cat_name}\")"
            )
            cc = ws.cell(row=ri, column=col, value=count_formula)
            cc.alignment = Alignment(horizontal="center", vertical="center")
            cc.font      = Font(bold=True, size=11)
            cc.border    = BOX
            if alt_fill:
                cc.fill = alt_fill

            # % Share — this category count / month total
            col_letter = get_column_letter(col)
            total_ref  = f"{col_letter}{total_row}"
            pct_formula = f"=IF({total_ref}>0,{col_letter}{ri}/{total_ref},0)"
            pc = ws.cell(row=ri, column=col + 1, value=pct_formula)
            pc.number_format = "0.0%"
            pc.alignment     = Alignment(horizontal="center", vertical="center")
            pc.font          = Font(size=10, color="666660")
            pc.border        = BOX
            if alt_fill:
                pc.fill = alt_fill

            col += 2

        ws.row_dimensions[ri].height = 20

    # ── Total row — SUM formulas ─────────────────────────────────────────
    tot_fill = PatternFill("solid", fgColor="E8EDF5")
    t1 = ws.cell(row=total_row, column=1, value="TOTAL")
    t1.font      = Font(bold=True, size=11)
    t1.fill      = tot_fill
    t1.border    = BOX
    t1.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    first_data = 4
    last_data  = first_data + len(ALL_CATS) - 1   # includes Other

    col = 2
    for month in all_months_data:
        col_letter  = get_column_letter(col)
        sum_formula = f"=SUM({col_letter}{first_data}:{col_letter}{last_data})"
        tc = ws.cell(row=total_row, column=col, value=sum_formula)
        tc.font      = Font(bold=True, size=12)
        tc.fill      = tot_fill
        tc.border    = BOX
        tc.alignment = Alignment(horizontal="center", vertical="center")

        # % column for the total row = 100 %
        pc = ws.cell(row=total_row, column=col + 1, value="100%")
        pc.font      = Font(size=10, color="666660")
        pc.fill      = tot_fill
        pc.border    = BOX
        pc.alignment = Alignment(horizontal="center", vertical="center")
        col += 2

    ws.row_dimensions[total_row].height = 24

    # ── Vertical clustered bar chart (openpyxl native API) ───────────────────
    if all_months_data:
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.legend import Legend

        # Chart rows: data rows only, no TOTAL row
        chart_last_row = 3 + len(ALL_CATS)

        chart          = BarChart()
        chart.type     = "col"
        chart.barDir   = "col"
        chart.grouping = "clustered"
        chart.overlap  = 0
        chart.title    = "Issue Count by Category — All Months"
        chart.height   = 18
        chart.width    = 30

        # Fix 1: correct axis positions for a vertical (column) chart
        chart.x_axis.axPos = "b"   # category axis at bottom
        chart.y_axis.axPos = "l"   # value axis on left
        chart.x_axis.title = None
        chart.y_axis.title = None

        # Fix 2: X-axis label rotation — -45 degrees via txPr/bodyPr
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.delete     = False
        try:
            from openpyxl.drawing.text import RichText, BodyProperties
            chart.x_axis.txPr = RichText(bodyPr=BodyProperties(rot=-2700000))
        except Exception:
            pass

        # Fix 3: Y-axis major gridlines
        try:
            from openpyxl.chart.axis import ChartLines
            chart.y_axis.majorGridlines = ChartLines()
        except Exception:
            pass

        chart.legend          = Legend()
        chart.legend.position = "b"
        chart.legend.overlay  = False

        cats_ref = Reference(ws, min_col=1, min_row=4, max_row=chart_last_row)
        chart.set_categories(cats_ref)

        month_colors = [
            "2E75B6",  # blue
            "ED7D31",  # orange
            "70AD47",  # green
            "FFC000",  # yellow
            "5B9BD5",  # light blue
            "FF0000",  # red
        ]

        col = 2
        for mi, month in enumerate(all_months_data):
            data_ref = Reference(ws, min_col=col, min_row=3,
                                 max_row=chart_last_row)
            chart.add_data(data_ref, titles_from_data=True)

            series = chart.series[mi]
            color  = month_colors[mi % len(month_colors)]
            try:
                series.graphicalProperties.solidFill      = color
                series.graphicalProperties.line.solidFill = color
            except Exception:
                pass

            series.dLbls               = DataLabelList()
            series.dLbls.showVal       = True
            series.dLbls.showCatName   = False
            series.dLbls.showSerName   = False
            series.dLbls.showLegendKey = False
            series.dLbls.showPercent   = False
            series.dLbls.dLblPos       = "outEnd"

            col += 2

        ws.add_chart(chart, f"A{total_row + 2}")

    # ── Column widths ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    col = 2
    for _ in all_months_data:
        ws.column_dimensions[get_column_letter(col)].width     = 13  # count
        ws.column_dimensions[get_column_letter(col + 1)].width = 10  # %
        col += 2

    ws.freeze_panes             = "B4"
    ws.sheet_view.showGridLines = False
    return ws
