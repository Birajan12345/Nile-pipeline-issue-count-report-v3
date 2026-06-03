"""
main.py
Multi-month orchestration:
  1. Auto-detect latest source file in input/
  2. Parse + categorize rows
  3. Load or create the permanent output Excel file
  4. Add new month tab (skip if month already exists)
  5. Refresh Cover Sheet  (always position 0, always latest month)
  6. Refresh All Months Summary  (position 1, reads every Raw sheet)
  7. Add Monthly Summary + Needs Review tabs for this month
  8. Save + write CSV archive + run log
"""
import csv
import logging
import os
import sys
from collections import Counter
from datetime import datetime

import openpyxl

from .config import (
    ALL_CATS, OUTPUT_DIR, OUTPUT_FILE,
    REPORT_MONTH_LABEL,
)
from .csv_reader import (
    detect_month_label, parse_rows,
    read_csv, read_html, read_xlsx,
)
from .excel_writer import (
    _short_label,
    write_all_months_summary,
    write_cover_sheet,
    write_raw_sheet,
    write_summary_sheet,
    write_uncategorized_sheet,
)
from .file_finder import find_latest_source


def _setup_logger(output_dir: str, label: str) -> logging.Logger:
    os.makedirs(output_dir, exist_ok=True)
    safe     = label.replace(" ", "_")
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(output_dir, f"run_{safe}_{ts}.log")

    logger = logging.getLogger("nile_tracker")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        "%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    ))
    logger.addHandler(fh)

    logger.info(f"Log file      : {log_path}")
    return logger


def read_source(input_path: str):
    """Dispatch to the correct reader based on file extension."""
    _, ext = os.path.splitext(input_path.lower())
    if ext in (".html", ".htm"):
        return read_html(input_path)
    if ext in (".xlsx", ".xls"):
        return read_xlsx(input_path)
    return read_csv(input_path)


def _collect_all_months(wb) -> list[dict]:
    """
    Read all 'Raw — *' sheets in the workbook.
    Returns list of {label, counts} sorted oldest → newest.
    """
    import calendar

    months = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Raw — "):
            continue
        label = sheet_name[len("Raw — "):]
        ws    = wb[sheet_name]

        # Find the Category column by scanning the header row
        cat_col = 22  # default (column V)
        for cell in ws[2]:
            if cell.value and str(cell.value).strip() == "Category":
                cat_col = cell.column
                break

        counts: dict[str, int] = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            val = row[cat_col - 1]
            if val:
                key = str(val)
                counts[key] = counts.get(key, 0) + 1

        months.append({"label": label, "counts": counts})

    def _sort_key(m):
        parts = m["label"].split()
        if len(parts) >= 2:
            try:
                month_num = list(calendar.month_abbr).index(
                    parts[0][:3].capitalize()
                )
                return int(parts[-1]) * 100 + month_num
            except Exception:
                pass
        return 0

    months.sort(key=_sort_key)
    return months


def _get_prev_month_counts(wb, current_raw_sheet: str) -> dict:
    """
    Read category counts from the most recent previous month's Raw sheet
    in the workbook (excluding the current month being written).
    Returns empty dict if no previous Raw sheet exists.
    """
    import calendar

    raw_sheets = [
        s for s in wb.sheetnames
        if s.startswith("Raw — ") and s != current_raw_sheet
    ]
    if not raw_sheets:
        return {}

    def _sort_key(name: str) -> int:
        label = name[len("Raw — "):]
        parts = label.split()
        if len(parts) >= 2:
            try:
                mn = list(calendar.month_abbr).index(
                    parts[0][:3].capitalize()
                )
                return int(parts[-1]) * 100 + mn
            except Exception:
                pass
        return 0

    raw_sheets.sort(key=_sort_key)
    prev_sheet_name = raw_sheets[-1]   # most recent previous month
    prev_ws = wb[prev_sheet_name]

    # Locate Category column from header row (row 2)
    headers = [cell.value for cell in list(
        prev_ws.iter_rows(min_row=2, max_row=2)
    )[0]]
    try:
        cat_idx = headers.index("Category")
    except ValueError:
        cat_idx = 21   # default column V (0-based index 21)

    counts: dict[str, int] = {}
    for row in prev_ws.iter_rows(min_row=3, values_only=True):
        val = row[cat_idx]
        if val:
            key = str(val)
            counts[key] = counts.get(key, 0) + 1

    print(f"  Prev month    : {prev_sheet_name}  ({sum(counts.values())} tickets)")
    return counts


def main(argv=None):
    argv = argv or sys.argv

    print(f"\n{'═' * 52}")
    print(f"  NILE Pipeline Issue Tracker  —  Multi-Month")
    print(f"{'═' * 52}")

    # ── 1. Resolve source file ─────────────────────────────────────────
    if len(argv) > 1:
        source_path = argv[1]
        print(f"\n  Source file   : {source_path}  (from argument)")
    else:
        source_path = find_latest_source()

    # ── 2. Parse + categorize ──────────────────────────────────────────
    raw_rows, _ = read_source(source_path)
    rows = parse_rows(raw_rows)

    # ── 3. Month label ─────────────────────────────────────────────────
    month_label    = REPORT_MONTH_LABEL or detect_month_label(rows)
    short_label    = _short_label(month_label)
    raw_sheet_name = f"Raw — {short_label}"
    summary_name   = f"Summary — {short_label}"
    review_name    = f"Needs Review — {short_label}"
    print(f"  Report month  : {month_label}  ({raw_sheet_name})")

    logger = _setup_logger(OUTPUT_DIR, month_label)
    logger.info(f"Source        : {source_path}")
    logger.info(f"Report month  : {month_label}")
    logger.info(f"Total rows    : {len(rows)}")

    # ── 4. Category counts ─────────────────────────────────────────────
    cat_counts = Counter(r["Category"] for r in rows)

    # ── 5. Load or create permanent workbook ──────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    xlsx_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)

    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        print(f"\n  Existing report: {OUTPUT_FILE}  (adding {month_label})")
        logger.info(f"Existing report loaded: {xlsx_path}")
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        print(f"\n  Creating new report: {OUTPUT_FILE}")
        logger.info(f"New report created: {xlsx_path}")

    # prev_counts — read directly from the previous Raw sheet in the workbook
    prev_counts = _get_prev_month_counts(wb, raw_sheet_name)
    if prev_counts:
        logger.info(f"Prev month    : {sum(prev_counts.values())} tickets (from workbook Raw sheet)")
    else:
        logger.info("Prev month    : none (first run or no previous Raw sheet)")

    # ── 6. Category breakdown ──────────────────────────────────────────
    print(f"\n  Category breakdown:")
    for name in ALL_CATS:
        if cat_counts[name]:
            diff  = cat_counts[name] - prev_counts.get(name, 0)
            trend = f"  ▲ +{diff}" if diff > 0 else (f"  ▼ {diff}" if diff < 0 else "")
            print(f"    {name:<36} {cat_counts[name]}{trend}")
            logger.info(f"  {name:<36} {cat_counts[name]}")
    print(f"    {'─' * 42}")
    print(f"    {'TOTAL':<36} {len(rows)}")

    # ── 7. Add Raw sheet (skip if month already exists) ────────────────
    if raw_sheet_name in wb.sheetnames:
        print(f"\n  WARNING: '{raw_sheet_name}' already exists — skipping write.")
        logger.warning(f"Sheet '{raw_sheet_name}' already exists — skipped.")
    else:
        raw_ws = wb.create_sheet(raw_sheet_name)
        write_raw_sheet(raw_ws, rows, month_label)
        raw_ws.title = raw_sheet_name   # override title set inside write_raw_sheet
        logger.info(f"Raw sheet added: {raw_sheet_name}")

    # ── 8. Refresh Cover Sheet (always position 0, latest month) ───────
    if "Cover Sheet" in wb.sheetnames:
        del wb["Cover Sheet"]
    cover_ws = wb.create_sheet("Cover Sheet", 0)
    write_cover_sheet(
        cover_ws, rows, month_label,
        dict(cat_counts), prev_counts,
        raw_sheet_name=raw_sheet_name,
    )
    logger.info("Cover Sheet refreshed")

    # ── 9. Refresh All Months Summary (position 1, after Cover Sheet) ──
    all_months = _collect_all_months(wb)
    write_all_months_summary(wb, all_months)
    logger.info(f"All Months Summary refreshed: {len(all_months)} months")

    # ── 10. Add / refresh Monthly Summary for this month ───────────────
    if summary_name in wb.sheetnames:
        del wb[summary_name]
    summary_ws = wb.create_sheet(summary_name)
    write_summary_sheet(
        summary_ws, rows, month_label,
        prev_counts=prev_counts,
        raw_sheet_name=raw_sheet_name,
    )
    summary_ws.title = summary_name     # override "Monthly Summary"
    logger.info(f"Summary sheet: {summary_name}")

    # ── 11. Add / refresh Needs Review for this month ──────────────────
    if review_name in wb.sheetnames:
        del wb[review_name]
    review_ws = wb.create_sheet(review_name)
    write_uncategorized_sheet(review_ws, rows, raw_sheet_name=raw_sheet_name)
    review_ws.title = review_name       # override "Needs Review"
    logger.info(f"Needs Review sheet: {review_name}")

    # ── 12. Refresh _categories hidden sheet ───────────────────────────
    if "_categories" in wb.sheetnames:
        del wb["_categories"]
    cat_ws = wb.create_sheet("_categories")
    for i, cat in enumerate(ALL_CATS, 1):
        cat_ws.cell(row=i, column=1, value=cat)
    cat_ws.sheet_state = "hidden"

    # ── 13. Reorder sheets into canonical order ────────────────────────
    #  Cover Sheet → All Months Summary → [Raw/Summary/Needs Review per month
    #  in chronological order] → _categories (hidden)
    canonical: list[str] = ["Cover Sheet", "All Months Summary"]
    for m in _collect_all_months(wb):          # already sorted oldest→newest
        lbl = m["label"]
        for prefix in ("Raw — ", "Summary — ", "Needs Review — "):
            sn = f"{prefix}{lbl}"
            if sn in wb.sheetnames:
                canonical.append(sn)
    if "_categories" in wb.sheetnames:
        canonical.append("_categories")
    # Any leftover sheets (shouldn't exist, but keep them at the end)
    for sn in wb.sheetnames:
        if sn not in canonical:
            canonical.append(sn)

    for target_pos, sn in enumerate(canonical):
        if sn in wb.sheetnames:
            current_pos = wb.sheetnames.index(sn)
            if current_pos != target_pos:
                wb.move_sheet(sn, offset=target_pos - current_pos)

    # ── 14. Save Excel ─────────────────────────────────────────────────
    wb.save(xlsx_path)
    logger.info(f"Excel saved   : {xlsx_path}")
    print(f"\n  ✓ Report saved  → {xlsx_path}")

    # ── 14. Write CSV archive for this month ───────────────────────────
    csv_name  = f"Nile_{short_label.replace(' ', '_')}_archive.csv"
    csv_path  = os.path.join(OUTPUT_DIR, csv_name)
    fieldnames = [
        "Key", "Issue id", "Parent id", "Summary", "Description",
        "Comment", "Close Notes", "Categories", "New Categories",
        "Assign", "Status", "Created", "Updated",
        "Custom field (Assignment Group)", "Assignee", "Reporter",
        "Resolution", "Custom field (Comments)",
        "Inward issue link (Cloners)", "Outward issue link (Cloners)",
        "Inward issue link (Relates)", "Category", "Review Required",
    ]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    logger.info(f"CSV archive   : {csv_path}")
    print(f"  ✓ CSV archive   → {csv_path}")

    # ── Sheet order summary ────────────────────────────────────────────
    print(f"\n  Sheets in workbook:")
    for name in wb.sheetnames:
        hidden = " (hidden)" if wb[name].sheet_state == "hidden" else ""
        print(f"    {name}{hidden}")

    print(f"\n{'═' * 52}\n")


if __name__ == "__main__":
    main()
